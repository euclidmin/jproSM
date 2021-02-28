import openpyxl
import sys
from datetime import datetime
import os



def funcname():
    return sys._getframe(1).f_code.co_name + "()"

def make_file_name():
    print(funcname())
    now = datetime.now()
    fname = 'JP_SM'+str(now.year)+'-'+str(now.month)+'-'+str(now.day)+'_'+str(now.hour)+str(now.minute)+str(now.second)+'.xlsx'
    return fname

def make_title_name():
    print(funcname())
    now = datetime.now()
    fname = '재고리스트'+str(now.year)+'-'+str(now.month)+'-'+str(now.day)+'_'+str(now.hour)+str(now.minute)+str(now.second)
    return fname


class BuildStock:
    def __init__(self):
        self.stock_master_file = None
        self.transaction_file = None
        self.stock_master_workbook = None
        self.transaction_workbook = None
        self.stock_list_from_transaction = None

        self.name_option_cells = None
        self.name_option_dict = {}
        self.stock_master_sheet_max_cnt = 0
        print(funcname())

    def load_workbook(self):
        print(funcname())

        def _open_transaction_file():
            print(funcname())
            return openpyxl.load_workbook('스마트스토어_주문조회_20210301_0715.xlsx')

        def _open_stock_master_file():
            print(funcname())
            if os.path.exists("JP_SM.xlsx") :
                stock_excel_file_name = "JP_SM.xlsx"
            else:
                stock_excel_file_name = make_file_name()
                wb = openpyxl.Workbook()
                wb.save(stock_excel_file_name)
            return openpyxl.load_workbook(stock_excel_file_name)

        self.stock_master_workbook = _open_stock_master_file()
        self.transaction_workbook = _open_transaction_file()


    def write_items_in_stock_master_file(self, items):
        print(funcname())
        sheet = self.stock_master_workbook.active
        sheet.title = "재고리스트"
        sheet.cell(row=1, column=1, value="판매품명")
        sheet.cell(row=1, column=2, value="재고개수")
        stock_master_item_value = self.get_stock_list_from_stock_master_file()

        cnt = 0
        for item_name in items.keys():
            if item_name not in stock_master_item_value:
                print(item_name)
                cnt += 1
                row_cnt = self.stock_master_sheet_max_cnt + cnt

                sheet.cell(row=row_cnt, column=1, value=item_name)
                sheet.cell(row=row_cnt, column=2, value=items[item_name])

        print(cnt)
        stock_excel_file_name = make_file_name()
        self.stock_master_workbook.save(stock_excel_file_name)




    def get_stock_list_from_transaction(self):
        print(funcname())
        def get_name_and_option_cells_from_transaction():
            print(funcname())
            transaction_worksheet = self.transaction_workbook['주문조회']
            max_cnt = transaction_worksheet.max_row
            ret = transaction_worksheet['G2':'H' + str(max_cnt)]
            # print(ret)
            return ret

        def make_name_and_option(cells):
            print(funcname())
            name_option_dict = {}
            for index in range(0, len(cells)):
                name = cells[index][0].value
                option = cells[index][1].value
                name_option = name + ' | ' + option
                # 재고 아이템별 중복 없이 dict에 초기값 0과 함께 저장
                name_option_dict[name_option] = 0
                # print(str(index) + ' ' + name_option)

            print(len(name_option_dict))
            return name_option_dict

        cells = get_name_and_option_cells_from_transaction()
        ret = make_name_and_option(cells)

        return ret


    def get_stock_list_from_stock_master_file(self):
        print(funcname())
        stock_master_file_worksheet = self.stock_master_workbook['재고리스트']
        max_cnt = stock_master_file_worksheet.max_row
        self.stock_master_sheet_max_cnt = max_cnt
        cells = stock_master_file_worksheet['A2':'B' + str(max_cnt)]
        # print(ret)
        item_name_value_dict = {}
        for index in range(0, len(cells)):
            item_name = cells[index][0].value
            value = cells[index][1].value
            # 재고 아이템별 중복 없이 dict에 초기값 0과 함께 저장
            item_name_value_dict[item_name] = value
            # print(str(index) + ' ' + name_option)
        return item_name_value_dict





class GoodsOut:
    def __init__(self):
        self.stock_master_workbook = None
        self.transaction_workbook = None
        self.stock_master_sheet = None
        self.transaction_sheet = None


    def load_workbook(self):
        print(funcname())

        def _open_transaction_file():
            print(funcname())
            return openpyxl.load_workbook('스마트스토어_주문조회_20210301_0715.xlsx')

        def _open_stock_master_file():
            print(funcname())
            if os.path.exists("JP_SM.xlsx") :
                stock_excel_file_name = "JP_SM.xlsx"
            else:
                print("JP_SM.xlsx 파일이 없습니다.")
            return openpyxl.load_workbook(stock_excel_file_name)

        self.stock_master_workbook = _open_stock_master_file()
        self.transaction_workbook = _open_transaction_file()
        self.stock_master_sheet = self.stock_master_workbook['재고리스트']
        self.transaction_sheet = self.transaction_workbook['주문조회']


    # def filter_out(self):
    #     max_cnt = self.transaction_sheet.max_row
    #     sheet = self.transaction_sheet
    #
    #     del_list = []
    #     for i in range(2, max_cnt):
    #         state = sheet.cell(row=i, column=4).value
    #         print(str(i)+' '+state)
    #         if state == '취소':
    #             del_list.append(i)
    #     print(del_list)
    #
    #     for idx in del_list :
    #         sheet.delete_rows(idx, amount=1)
    #         print(idx)

    #
    # def save(self):
    #     file_name = make_file_name()
    #     self.transaction_workbook.save(file_name)




    def get_order_list_from_transaction(self):
        print(funcname())
        def get_name_and_option_cnt_cells_from_transaction():
            print(funcname())
            transaction_worksheet = self.transaction_sheet
            max_cnt = transaction_worksheet.max_row
            ret = transaction_worksheet['D2':'I' + str(max_cnt)]
            # print(ret)
            return ret

        def make_name_and_option_cnt(cells):
            print(funcname())
            name_option_dict = {}
            for index in range(0, len(cells)):
                state = cells[index][0].value
                name = cells[index][3].value
                option = cells[index][4].value
                count = cells[index][5].value
                name_option = name + ' | ' + option
                # 재고 아이템별 중복 없이 dict에 주문 수량(출고 수량) 저장
                # 취소된 주문은 제외한다.
                if state != '취소' :
                    name_option_dict[name_option] = count
                # print(str(index) + ' ' + name_option)

            print(len(name_option_dict))
            return name_option_dict

        cells = get_name_and_option_cnt_cells_from_transaction()
        ret = make_name_and_option_cnt(cells)

        return ret

    def update_stock_master_file(self, order_list):
        wb = self.stock_master_workbook
        sheet = wb.create_sheet()
        sheet.title = make_title_name()
        sheet.cell(row=1, column=1, value="판매품명")
        sheet.cell(row=1, column=2, value="재고개수")

        stock_master_item_value = self.get_stock_list_from_stock_master_file()

        for order in order_list.keys():
            stock_master_item_value[order] -= order_list[order]

        for i, item in enumerate(stock_master_item_value.keys()):
            sheet.cell(row=(i+2), column=1, value=item)
            sheet.cell(row=(i+2), column=2, value=stock_master_item_value[item])

        stock_excel_file_name = make_file_name()
        wb.save(stock_excel_file_name)


    def get_stock_list_from_stock_master_file(self):
        print(funcname())
        stock_master_file_worksheet = self.stock_master_workbook['재고리스트']
        max_cnt = stock_master_file_worksheet.max_row
        self.stock_master_sheet_max_cnt = max_cnt
        cells = stock_master_file_worksheet['A2':'B' + str(max_cnt)]
        # print(ret)
        item_name_value_dict = {}
        for index in range(0, len(cells)):
            item_name = cells[index][0].value
            value = cells[index][1].value
            # 재고 아이템별 중복 없이 dict에 초기값 0과 함께 저장
            item_name_value_dict[item_name] = value
            # print(str(index) + ' ' + name_option)
        return item_name_value_dict