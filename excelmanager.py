import openpyxl
from datetime import datetime
# import sys



class ExcelManager:
    def __init__(self):
        self.workbook = None
        self.p_sheet = None          # Product sheet
        self.t_sheet = None          # Transaction sheet
        self.m_sheet = None          # Manufacturer sheet
        self.product_id = None
        self.type = None
        self.count = None

        self.selected = {}                    # product : 다미끼 , weight : 250 , color : #01, size : #3
        

    def open_excel(self):
        self.workbook  = openpyxl.load_workbook('jiggingpro.xlsx')
        # self.t_sheet = self.workbook.get_sheet_by_name('Transaction')
        # self.p_sheet = self.workbook.get_sheet_by_name('Product')
        self.t_sheet = self.workbook['Transaction']
        self.p_sheet = self.workbook['Product']
        self.m_sheet = self.workbook['Manufacturer']



    def _update_count(self, row, cnt):
        self.p_sheet['G' + str(row)] = self.p_sheet['G' + str(row)].value + cnt

    def _find_row_by_product_ID(self, id):
        ret = None
        for i in range(2, self.p_sheet.max_row+1):
            if self.p_sheet['A'+str(i)].value == id :
                ret = i
                break
            else :
                pass
        return ret


    def update_stock(self):
        get_product_ID = lambda row : self.t_sheet['C'+str(row)].value
        get_count = lambda row : self.t_sheet['D'+str(row)].value
        get_type = lambda row : self.t_sheet['B'+str(row)].value
        p_n = lambda t : 1 if t == 'buy' else -1

        for row in range(2, self.t_sheet.max_row+1):
            id = get_product_ID(row)
            type = get_type(row)
            cnt = get_count(row)
            cnt = cnt * p_n(type)

            p_row = self._find_row_by_product_ID(id)
            self._update_count(p_row, cnt)

    def save(self):
        now = datetime.now()
        fname = 'jiggingpro'+str(now.year)+'-'+str(now.month)+'-'+str(now.day)+'_'+str(now.hour)+str(now.minute)+str(now.second)+'.xlsx'
        self.workbook.save(fname)

    def get_manufacturer(self):
        manuf_cells = lambda colm: self.m_sheet[colm]
        cell_value = lambda cell : cell.value

        manufacturer_set = set(map(cell_value, manuf_cells('B')))
        manufacturer_set.remove('NAME')      # 첫줄의 카테고리 이름을 제외 시킨다.
        return list(manufacturer_set)

    def get_products(self, name):            # name 은 제조사 이름 ex)다미끼
        self.selected['manufacturer'] = name
        product_sheet = self.workbook[name]

        product_cells = lambda colm: product_sheet[colm]
        cell_value = lambda cell: cell.value

        product_set = set(map(cell_value, product_cells('B')))
        product_set.remove('NAME')
        return list(product_set)

    def get_weight(self, name):               # name 은 제조사 이름 ex)랜스롱
        self.selected['product'] = name
        manuf_name = self.selected['manufacturer']
        product_sheet = self.workbook[manuf_name]

        get_product_name = lambda row : product_sheet['B'+str(row)].value
        get_weight = lambda row : product_sheet['C'+str(row)].value
        weight_set = set()
        for row in range(2, product_sheet.max_row+1):
            product_name = get_product_name(row)
            if product_name == name :
                weight_set.add(get_weight(row))
        return list(weight_set)



